#!/usr/bin/env python
# -*- coding:utf-8 -*-


import re
import urllib.error
import urllib.request as urq
from datetime import datetime
from threading import Thread, Semaphore

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font
from youtube_api import YouTubeDataAPI


#全局变量
vtuber_names = {
	'0': 'ときのそら', '1': 'ロボ子さん', '2': 'さくらみこ', '3': '夜空メル',
	'4': '白上フブキ', '5': '夏色まつり', '6': '赤井はあと',
	'7': 'アキ・ローゼンタール', '8': '大空スバル','9': '湊あくあ',
	'10': '紫咲シオン', '11': '百鬼あやめ','12': '癒月ちょこ', '13': '大神ミオ',
	'14': '猫又おかゆ', '15': '戌神ころね', '16': '兎田ぺこら',
	'17': '潤羽るしあ', '18': '白銀ノエル', '19': '宝鐘マリン',
	'20': '不知火フレア', '21': '星街すいせい', '22': 'AZKi', '23': '天音かなた',
	'24': '常闇トワ', '25': '角巻わため', '26': '桐生ココ', '27': '姫森ルーナ'
}


# 工作表抬头
def title_excel(sheet):
	global vtuber_names
	sheet.freeze_panes = 'B2'
	sheet.column_dimensions['A'].width = 18.50
	sheet.column_dimensions['B'].width = 18.50
	for fc in range(len(vtuber_names)):
		sheet.cell(row=fc+2, column=1).value = vtuber_names[str(fc)]
		sheet.cell(row=fc+2, column=1).font = Font(bold=True)
	return sheet


# 访问单个网页
def ask_url(url):
	# Shadowsocks代理访问
	proxies = {
		'https': 'https://127.0.0.1:1080',
		'http': 'http://127.0.0.1:1080'
	}
	headers = {
		'user-agent':('Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit'
				'/537.36 (KHTML, like Gecko) Chrome'
				'/83.0.4103.61 Safari/537.36')
	}
	opener = urq.build_opener(urq.ProxyHandler(proxies))
	urq.install_opener(opener)
	# 请求与响应
	try:
		req = urq.Request(url, headers=headers)
		response = urq.urlopen(req, timeout=5)
	except urllib.error.URLError as e:
		if hasattr(e, 'code'):
			print(e.code)
		if hasattr(e, 'reason'):
			print(e.reason)
		return e.code
	else:
		html = response.read().decode('utf-8')
		return html


# 线程：nico本家
class NiconicoDataT(Thread):

	def __init__(self, user, index, sheet):
		super().__init__()
		self.user = user
		self.index = index
		self.sheet = sheet
	
	def run(self):
		url = 'https://www.nicovideo.jp/user/' + self.user
		html = ask_url(url)
		if isinstance(html, int) and len(str(html)) == 3:
			print('HTMLError!')
			return None
		# 解析数据
		bs = BeautifulSoup(html, 'html.parser')
		for i in bs.find_all('div', class_='profile'):
			i = str(i)
			# 频道名, Follow, Follower, Stamp, Level, EXP
			name = re.findall(r'<h2>(.*?)<small>さん</small></h2>', i)
			ffs = re.findall(r'<span class="num">(.*?)</span>', i)
			lv = re.findall(r'<a class="user-level-num" data-user-level-link="">(.*?)</a>', i)
			exp = re.findall(r'<div class="user-level-meter-container" data-title="(.*?)">', i)
		channel = name + ffs + lv + exp
		for col in range(2, 8):
			self.sheet.cell(row=self.index+2, column=col).value = channel[col-2]
			if col == 4:
				self.sheet.cell(row=self.index+2, column=col).font = Font(color='FF0000')

	def get_result(self):
		return self.sheet


# 数据获取：nico本家
def get_niconico(sheet):
	row_title = ['Row', 'ChName', 'Follow', 'Follower', 'Stamp', 'Level', 'EXP']
	for rt in range(2, 8):
		sheet.cell(row=1, column=rt).value = row_title[rt-1]
		sheet.cell(row=1, column=rt).font = Font(bold=True)
	users_niconico = [
		'4103372', '25022047', '40568633', '14752518', '2929601', '52019736',
		'', '', '', '45194653', '', '', '', '15832191', '30234055', '30685941',
		'48021351', '11827865', '7190326', '2182682', '361972', '29269692',
		'1279558', '26272699', '25440222', '11949668', '56821510', '36217532'
	]
	threads_niconico = []
	for index, user in enumerate(users_niconico):
		if user:
			t = NiconicoDataT(user, index, sheet)
			t.start()
			threads_niconico.append(t)
			sheet = t.get_result()
	for th in threads_niconico:
		th.join()
	return sheet


# 线程：nico社区
class NicocommuDataT(Thread):

	def __init__(self, user, index, sheet):
		super().__init__()
		self.user = user
		self.index = index
		self.sheet = sheet
	
	def run(self):
		url = 'https://com.nicovideo.jp/community/co' + self.user
		html = ask_url(url)
		if isinstance(html, int) and len(str(html)) == 3:
			print('HTMLError!')
			return None
		bs = BeautifulSoup(html, 'html.parser')
		for i in bs.find_all('div', class_='communityData'):
			i = str(i)
			# 频道名
			name = re.findall(r'<a href="/community/co\d*">[\n\t]*(.*?)[\n\t]*</a>', i, re.S)
		for j in bs.find_all('div', class_='communityRegist_text'):
			j = str(j)
			# Level, Follower, 最大ff数，需要升级人数
			lv = re.findall(r'<dd class="content">(.*?)</dd>', j)
			ffs = re.findall(r'<dd class="content memberCount">.*?(\d*?)<span class="unit">人</span>', j, re.S)
			maxs = re.findall(r'<span class="max">[(]最大:(\d*?)人[)]</span>', j)
			needs = re.findall(r'があと<span class="num">(\d*?)</span>人でレベル', j)
		channel = name + lv + ffs + maxs + needs
		for col in range(2, 7):
			self.sheet.cell(row=self.index+2, column=col).value = channel[col-2]
			if col == 4:
				self.sheet.cell(row=self.index+2, column=col).font = Font(color='FF0000')

	def get_result(self):
		return self.sheet


# 数据获取：nico社区
def get_nicocommu(sheet):
	row_title = ['Row', 'ChName', 'Level', 'Follower', 'MaxFFer', 'NeedFFer']
	for rt in range(2, 7):
		sheet.cell(row=1, column=rt).value = row_title[rt-1]
		sheet.cell(row=1, column=rt).font = Font(bold=True)
	users_nicocommu = [
		'2870964', '', '', '370519', '423126', '3513608', '', '', '',
		'2665206', '', '', '', '112548', '', '1881807', '2805474', '106424',
		'19577', '357317', '50850', '', '2440266', '2491317', '1860553',
		'246853', '3340513', '2963469'
	]
	threads_nicocommu = []
	for index, user in enumerate(users_nicocommu):
		if user:
			t = NicocommuDataT(user, index, sheet)
			t.start()
			threads_nicocommu.append(t)
			sheet = t.get_result()
	for th in threads_nicocommu:
		th.join()
	return sheet


# 线程：twitter
class TwitterDataT(Thread):

	def __init__(self, user, index, sem, sheet):
		super().__init__()
		self.user = user
		self.index = index
		self.sem = sem
		self.sheet = sheet
	
	def run(self):
		url = 'https://twilog.org/' + self.user		# 没有key，不想写小作文
		html = ask_url(url)
		if isinstance(html, int) and len(str(html)) == 3:
			print('HTMLError!')
			return None
		# 解析数据
		bs = BeautifulSoup(html, 'html.parser')
		for i in bs.find_all('div', id='user-info-content'):
			i = str(i)
			# 频道名, Follow, Follower, 备注
			name = re.findall(r'<strong>(.*?)</strong><span>', i)
			fo = re.findall(r'target="_blank">(.*?)</a>フォロー</li>', i)
			ffs = re.findall(r'target="_blank">(.*?)</a>フォロワー</li>', i)
			com = re.findall(r'<li class="user-info-comment">(.*?)</li>', i, re.S)[0]
			if '<' in com: 
				com = re.sub(r'<.*?>', '', com)
			if ('\n' or '\t') in com:
				com = re.sub(r'[\n\t]', '', com)
		channel = name + fo + ffs + [com]
		for col in range(2, 6):
			self.sheet.cell(row=self.index+2, column=col).value = channel[col-2]
			if col == 4:
				self.sheet.cell(row=self.index+2, column=col).font = Font(color='FF0000')
		self.sem.release()

	def get_result(self):
		return self.sheet


# 数据获取：twitter
def get_twitter(sheet):
	row_title = ['Row', 'Twiname', 'Follow', 'Follower', 'Comments']
	for rt in range(2, 6):
		sheet.cell(row=1, column=rt).value = row_title[rt-1]
		sheet.cell(row=1, column=rt).font = Font(bold=True)
	users_twitter = [
		'', '', '', 'kuroneko_datenn', '', 'masimasi0',
		'pikurusu_003', 'block', 'mecchababumi', 'rin_co_co', 'block', 'block', 
		'whatweneed8837', 'meianmei', 'IKAIKAMAKKURO', 'mysk2525', 'block',
		'block', 'canan8181', '', 'dankuri', 'Endolu', '7utauta', 'Nozo9n',
		'block', 'miketama27', '_ka_son', 'kanonchan1203'
	]
	threads_twitter = []	
	sem = Semaphore(10)			# 线程锁
	for index, user in enumerate(users_twitter):
		if user == 'block':
			sheet.cell(row=index+2, column=2).value = 'BLOCKED'
			sheet.cell(row=index+2, column=2).font = Font(color='808080')
		elif user:
			sem.acquire()
			t = TwitterDataT(user, index, sem, sheet)
			t.start()
			threads_twitter.append(t)
			sheet = t.get_result()
	for th in threads_twitter:
		th.join()
	return sheet


# 线程：youtube
class YoutubeDataT(Thread):

	def __init__(self, user, index, sheet):
		super().__init__()
		self.user = user
		self.index = index
		self.sheet = sheet
	
	def run(self):
		# 记得开全局
		api_key = 'AIzaSyDdwxeNNrbIQomjOmfnpFwWyRA5b9zHdNQ'
		yt = YouTubeDataAPI(api_key)
		dir = yt.get_channel_metadata(self.user)
		# 频道名，关注，播放数，简述
		channel = [dir['title'], dir['subscription_count'], dir['view_count'],
			 re.sub('\n', ' ', dir['description'])]
		for col in range(2, 6):
			# 千位分隔符
			if channel[col-2].isdigit():
				channel[col-2] = format(int(channel[col-2]), ',')
			self.sheet.cell(row=self.index+2, column=col).value = channel[col-2]
			if col == 3:
				self.sheet.cell(row=self.index+2, column=col).font = Font(color='FF0000')
		# 隐藏值处理
		for x in range(len(channel)):
			if channel[x] == '0':
				self.sheet.cell(row=self.index+2, column=x+2).value = 'HIDDEN'
				self.sheet.cell(row=self.index+2, column=x+2).font = Font(color='808080')

	def get_result(self):
		return self.sheet


# 数据获取：youtube
def get_youtube(sheet):
	row_title = ['Row', 'ChName', 'Subscribers', 'Views', 'Description']
	for rt in range(2, 6):
		sheet.cell(row=1, column=rt).value = row_title[rt-1]
		sheet.cell(row=1, column=rt).font = Font(bold=True)
	# 防止访问失败导致错误
	try:
		utest = urq.urlopen('https://www.youtube.com/', timeout=5)
	except:
		print('HTTPError: 404 Not Found.')
		return None
	users_youtube = [
		'', 'UCAV0LixnS-1Hk_bh5Vl6Isw', '', 'UCz4jhqrCfthF8NnldZeK_rw', '',
		'UChXPBr_T090MUkau4P8AF3g', 'UCxC0vx_abu7WtJfAnUwdMng', '',
		'UC3-xPUs13J3za-RZ-9qyAZA', 'UCBbGcCpI1NmpNdEV1qPQttw', '', '',
		'UCJd7dGbOFMI3tnr5KL4zEEQ', 'UC673c1ErzOqfHZv6J4HWZoA',
		'UCE6KXP8MqqqrD8alHkraq_Q', '', 'UCMbWhpVislQI7czEUUajP_Q',
		'UCCvGFp4-od62TmUCRze_alw', 'UCX8fFzD1Kdi8uCSmbOce08w',
		'UCWYJ0nzZqGZfyiUiVT2Z4MA', '', '', 'UCmhtmUBjkXOAetnaDq-XJ1g',
		'UCZHjpsV_0jcmkFvAwL1bkkw', 'UCkN4o220d-uc_S9b3i5-VOA',
		'UC8U-nKfWmJ6aEjPVeRFQVew', 'UC9ruVYPv7yJmV0Rh0NKA-Lw',
		'UCCmG1Rl23NQ63dMHsAjG38A'
	]
	threads_youtube = []
	for index, user in enumerate(users_youtube):
		if user:
			t = YoutubeDataT(user, index, sheet)
			t.start()
			threads_youtube.append(t)
			sheet = t.get_result()
	for th in threads_youtube:
		th.join()
	return sheet


# 线程：twitcasting
class TwicasDataT(Thread):

	def __init__(self, user, index, sheet):
		super().__init__()
		self.user = user
		self.index = index
		self.sheet = sheet
	
	def run(self):
		url = 'https://twitcasting.tv/' + self.user
		html = ask_url(url)
		if isinstance(html, int) and len(str(html)) == 3:
			print('HTMLError!')
			return None
		bs = BeautifulSoup(html, 'html.parser')
		for i in bs.find_all('div', class_='tw-user-nav-user'):
			i = str(i)
			# 频道名，Level, follower
			name = re.findall(r'<span class="tw-user-nav-name">(.*?)</span>', i)
			name = [name[0].strip()]
			lv = re.findall(r'class="tw-user-nav-stat-value">(\d*?)</span></li>', i)
			ffs = format(int(lv[-1]), ',')
		for j in bs.find_all('div', class_='tw-live-author__comment'):
			j = str(j)
			# 简介
			com = re.findall(r'<div class="comment" id="authorcomment">(.*?)</div>', j, re.S)[0]
			com = (re.sub(r'<(.*?)>|[\n\t]', '', j)).strip()
		channel = name + [lv[0], ffs, com]
		for col in range(2, 6):
			self.sheet.cell(row=self.index+2, column=col).value = channel[col-2]
			if col == 4:
				self.sheet.cell(row=self.index+2, column=col).font = Font(color='FF0000')

	def get_result(self):
		return self.sheet


# 数据获取：twicasting
def get_twicas(sheet):
	row_title = ['Row', 'Chname', 'Level', 'Follower', 'Comments']
	for rt in range(2, 6):
		sheet.cell(row=1, column=rt).value = row_title[rt-1]
		sheet.cell(row=1, column=rt).font = Font(bold=True)
	users_twicas = [
		'', 'hnmr_doll', '__nyan_xx', 'kuroneko_datenn', '', 'masimasi0',
		'pikurusu_003', 'miyukiii_lispon', 'mecchababumi', 'rin_co_co',
		'c:rinakuro0603', 'enm_a', 'whatweneed8837', '', 'IKAIKAMAKKURO',
		'mysk2525','', 'dicekawaiiiiii', 'canan8181', 'misaman339', 'dankuri',
		'', '7utauta', 'Nozo9n', 'NNNrm28', 'miketama27', '', 'kanonchan1203'
	]
	threads_twicas = []	
	for index, user in enumerate(users_twicas):
		if user:
			t = TwicasDataT(user, index, sheet)
			t.start()
			threads_twicas.append(t)
			sheet = t.get_result()
	for th in threads_twicas:
		th.join()
	return sheet


def main():
	t1 = datetime.now()
	wb = Workbook()
	ws = wb.active
	ws.title = 'Information'
	ws['A1'] = 'CopyRight:'
	ws['A2'] = 'Author: xiluo_shiro'
	ws['A3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

	ws1 = wb.create_sheet('Niconico')
	ws1 = title_excel(ws1)
	ws1 = get_niconico(ws1)
	print('Niconico done.')

	ws2 = wb.create_sheet('Nicocommu')
	ws2 = title_excel(ws2)
	ws2 = get_nicocommu(ws2)
	print('Nicocommu done.')
	
	ws3 = wb.create_sheet('Twitter')
	ws3 = title_excel(ws3)
	ws3 = get_twitter(ws3)
	print('Twitter done.')

	ws4 = wb.create_sheet('Youtube')
	ws4 = title_excel(ws4)
	ws4 = get_youtube(ws4)
	print('Youtube done.')

	ws5 = wb.create_sheet('Twitcasting')
	ws5 = title_excel(ws5)
	ws5 = get_twicas(ws5)
	print('Twitcasting done.')

	date = datetime.now().strftime('%Y-%m-%d')
	wb.save(date + '.xlsx')
	print('All done.')
	print('Time:', datetime.now()-t1)


if __name__ == '__main__':
	main()